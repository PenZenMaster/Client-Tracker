r"""
Module/Script Name: create_project_tracker.py
Path: E:\projects\Project Tracking\create_project_tracker.py

Description:
Excel project tracking workbook generator. Creates task management spreadsheet with
status tracking, data validation, summary metrics, and visualization charts.

Author(s):
Rank Rocket Co (C) Copyright 2025 - All Rights Reserved

Created Date:
2024-01-15

Last Modified Date:
2025-10-23

Version:
v1.02

License:
CC BY-SA 4.0 - https://creativecommons.org/licenses/by-sa/4.0/

Comments:
* v1.02 - Added type hints and improved documentation
* v1.01 - Added standardized file header
* v1.00 - Initial release with openpyxl project tracker

Module Functions:
    Creates an Excel workbook (ProjectTracker.xlsx) with three sheets:
    1. Tasks - Main task list with status dropdown validation
    2. Lists - Data validation source for status values
    3. Summary - Metrics dashboard with completion percentage and pie chart

Usage:
    Run directly to create ProjectTracker.xlsx in current directory:
    $ python create_project_tracker.py
"""

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore[import-untyped]
from openpyxl.chart import PieChart, Reference  # type: ignore[import-untyped]

# Initialize workbook
wb = Workbook()
ws_tasks = wb.active
ws_tasks.title = "Tasks"

headers = ["Task", "Subtasks", "Timeline", "Status", "Notes"]
ws_tasks.append(headers)

tasks_data = [
    ["Cloud Stacks", "2 initial stacks; 1 monthly ongoing", "Initial, Monthly", "", ""],
    ["Google Stacks", "4 per month", "Monthly", "", ""],
    [
        "CTR - Branded Keywords",
        "Set up tracking for 10 keywords",
        "Months 1-3",
        "",
        "Record monthly CTR data",
    ],
    [
        "CTR - Low-Competition Keywords",
        "Identify & track 10 keywords",
        "Monthly",
        "",
        "",
    ],
    [
        "CTR - High-Volume Keywords",
        "Month 1: 3 keywords; add 5 monthly up to 30",
        "Month 1 onward",
        "",
        "",
    ],
    ["Backlink Building", "Research, outreach, follow-up", "Ongoing", "", ""],
    ["Enhanced Schema", "Audit, plan, implement, test", "One-Time Project", "", ""],
]
for row in tasks_data:
    ws_tasks.append(row)

ws_lists = wb.create_sheet(title="Lists")
ws_lists["A1"] = "Not Started"
ws_lists["A2"] = "In Progress"
ws_lists["A3"] = "Completed"

dv = DataValidation(type="list", formula1="=Lists!$A$1:$A$3", allow_blank=True)
dv.error = "Please select a valid status"
dv.errorTitle = "Invalid Status"
ws_tasks.add_data_validation(dv)
dv.add("D2:D8")

ws_summary = wb.create_sheet(title="Summary")
ws_summary["A1"] = "Metric"
ws_summary["B1"] = "Value"
ws_summary["A2"] = "Total Tasks"
ws_summary["B2"] = "=COUNTA(Tasks!A2:A8)"
ws_summary["A3"] = "Completed Tasks"
ws_summary["B3"] = '=COUNTIF(Tasks!D2:D8, "Completed")'
ws_summary["A4"] = "Overall Completion (%)"
ws_summary["B4"] = "=IF(B2=0, 0, B3/B2)"
ws_summary["B4"].number_format = "0.00%"

ws_summary["A6"] = "Category"
ws_summary["B6"] = "Count"
ws_summary["A7"] = "Completed"
ws_summary["B7"] = '=COUNTIF(Tasks!D2:D8, "Completed")'
ws_summary["A8"] = "Not Completed"
ws_summary["B8"] = '=COUNTA(Tasks!D2:D8)-COUNTIF(Tasks!D2:D8, "Completed")'

pie = PieChart()
data = Reference(ws_summary, min_col=2, min_row=7, max_row=8)
labels = Reference(ws_summary, min_col=1, min_row=7, max_row=8)
pie.add_data(data, titles_from_data=False)
pie.set_categories(labels)
pie.title = "Project Completion Breakdown"
ws_summary.add_chart(pie, "D6")

output_filename = "ProjectTracker.xlsx"
wb.save(output_filename)
print(f"Excel workbook '{output_filename}' created successfully!")
